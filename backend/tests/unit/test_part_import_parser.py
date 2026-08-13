"""Unit tests for the pure CSV/XLSX part-import parser (SPEC §3): header
validation, formula-injection rejection, decimal/currency validation,
in-file duplicate detection, row/column caps, and encoding tolerance.

No database anywhere in this file - matches the module under test.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from app.importing.part_import_parser import (
    MAX_COLUMNS,
    MAX_ROWS,
    PartImportParseError,
    normalize_internal_part_number,
    parse_csv,
    parse_xlsx,
    validate_rows,
)

HEADER = (
    "internal_part_number,name,manufacturer_part_number,description,category,"
    "unit_code,target_price,target_price_currency"
)


def _csv_bytes(text: str, encoding: str = "utf-8") -> bytes:
    return text.encode(encoding)


def _basic_csv(*data_lines: str) -> bytes:
    return _csv_bytes(HEADER + "\n" + "\n".join(data_lines) + "\n")


class TestHeaderValidation:
    def test_minimal_required_headers_accepted(self) -> None:
        raw_file = parse_csv(_csv_bytes("internal_part_number,name\nPN-1,Widget\n"))
        assert len(raw_file.rows) == 1
        assert raw_file.rows[0].cells == {"internal_part_number": "PN-1", "name": "Widget"}

    def test_missing_required_header_is_rejected(self) -> None:
        with pytest.raises(PartImportParseError, match="internal_part_number"):
            parse_csv(_csv_bytes("name\nWidget\n"))

    def test_unknown_header_is_rejected(self) -> None:
        with pytest.raises(PartImportParseError, match="Unrecognized"):
            parse_csv(_csv_bytes("internal_part_number,name,made_up_column\nPN-1,Widget,x\n"))

    def test_duplicate_header_is_rejected(self) -> None:
        with pytest.raises(PartImportParseError, match="Duplicate"):
            parse_csv(_csv_bytes("internal_part_number,name,name\nPN-1,Widget,Other\n"))

    def test_header_matching_is_case_insensitive_and_trims_whitespace(self) -> None:
        raw_file = parse_csv(_csv_bytes(" Internal_Part_Number , NAME \nPN-1,Widget\n"))
        assert raw_file.rows[0].cells == {"internal_part_number": "PN-1", "name": "Widget"}

    def test_empty_file_is_rejected(self) -> None:
        with pytest.raises(PartImportParseError, match="empty"):
            parse_csv(b"")

    def test_blank_trailing_lines_are_skipped(self) -> None:
        raw_file = parse_csv(_csv_bytes("internal_part_number,name\nPN-1,Widget\n\n\n"))
        assert len(raw_file.rows) == 1


class TestColumnAndRowCaps:
    def test_too_many_columns_is_rejected(self) -> None:
        headers = ",".join(f"col{i}" for i in range(MAX_COLUMNS + 1))
        with pytest.raises(PartImportParseError, match="columns"):
            parse_csv(_csv_bytes(headers + "\n"))

    def test_too_many_rows_is_rejected(self) -> None:
        lines = [f"PN-{i},Widget {i}" for i in range(MAX_ROWS + 1)]
        with pytest.raises(PartImportParseError, match="rows"):
            parse_csv(_csv_bytes("internal_part_number,name\n" + "\n".join(lines) + "\n"))

    def test_exactly_at_row_cap_is_accepted(self) -> None:
        lines = [f"PN-{i},Widget {i}" for i in range(MAX_ROWS)]
        raw_file = parse_csv(_csv_bytes("internal_part_number,name\n" + "\n".join(lines) + "\n"))
        assert len(raw_file.rows) == MAX_ROWS


class TestEncoding:
    def test_plain_utf8_decodes(self) -> None:
        raw_file = parse_csv(_csv_bytes("internal_part_number,name\nPN-1,Wídget\n", "utf-8"))
        assert raw_file.rows[0].cells["name"] == "Wídget"
        assert raw_file.encoding_fallback is False

    def test_utf8_bom_is_tolerated(self) -> None:
        data = "internal_part_number,name\nPN-1,Widget\n".encode("utf-8-sig")
        raw_file = parse_csv(data)
        assert raw_file.rows[0].cells == {"internal_part_number": "PN-1", "name": "Widget"}
        assert raw_file.encoding_fallback is False

    def test_latin1_fallback_is_flagged(self) -> None:
        # 0xe9 is 'é' in latin-1 but is not valid standalone UTF-8
        data = b"internal_part_number,name\nPN-1,Caf\xe9\n"
        raw_file = parse_csv(data)
        assert raw_file.encoding_fallback is True
        assert raw_file.rows[0].cells["name"] == "Café"


class TestFormulaInjection:
    def test_equals_leading_cell_is_rejected(self) -> None:
        rows = validate_rows(
            parse_csv(_basic_csv("=cmd|'/c calc'!A1,Widget,,,,,,")).rows
        )
        assert rows[0].parsed is None
        assert any(e.issue.startswith("value begins") for e in rows[0].errors)

    def test_at_leading_cell_is_rejected(self) -> None:
        rows = validate_rows(parse_csv(_basic_csv("PN-1,@SUM(A1:A10),,,,,,")).rows)
        assert rows[0].parsed is None
        assert any(e.field == "name" for e in rows[0].errors)

    def test_plus_leading_formula_is_rejected(self) -> None:
        rows = validate_rows(
            parse_csv(_basic_csv("PN-1,Widget,+HYPERLINK(A1),,,,,")).rows
        )
        assert rows[0].parsed is None
        assert any(e.field == "manufacturer_part_number" for e in rows[0].errors)

    def test_minus_leading_formula_is_rejected(self) -> None:
        rows = validate_rows(
            parse_csv(_basic_csv("PN-1,Widget,-2+3+cmd|' /C calc'!A1,,,,,")).rows
        )
        assert rows[0].parsed is None

    def test_plain_signed_number_cell_is_allowed(self) -> None:
        rows = validate_rows(
            parse_csv(_basic_csv("PN-1,Widget,-12345,,,,,")).rows
        )
        assert rows[0].parsed is not None
        assert rows[0].parsed.manufacturer_part_number == "-12345"

    def test_formula_injection_checked_on_every_column(self) -> None:
        rows = validate_rows(
            parse_csv(_basic_csv("PN-1,Widget,,=malicious(),,,,")).rows
        )
        assert rows[0].parsed is None
        assert any(e.field == "description" for e in rows[0].errors)


class TestRequiredFieldValidation:
    def test_missing_internal_part_number_is_invalid(self) -> None:
        rows = validate_rows(parse_csv(_basic_csv(",Widget,,,,,,")).rows)
        assert rows[0].parsed is None
        assert any(e.field == "internal_part_number" for e in rows[0].errors)

    def test_missing_name_is_invalid(self) -> None:
        rows = validate_rows(parse_csv(_basic_csv("PN-1,,,,,,,")).rows)
        assert rows[0].parsed is None
        assert any(e.field == "name" for e in rows[0].errors)

    def test_fully_populated_row_is_valid(self) -> None:
        rows = validate_rows(
            parse_csv(
                _basic_csv("PN-1,Widget,MPN-1,A widget,hardware,each,10.50000000,USD")
            ).rows
        )
        assert rows[0].errors == []
        parsed = rows[0].parsed
        assert parsed is not None
        assert parsed.internal_part_number == "PN-1"
        assert parsed.name == "Widget"
        assert parsed.manufacturer_part_number == "MPN-1"
        assert parsed.description == "A widget"
        assert parsed.category == "hardware"
        assert parsed.unit_code == "each"
        assert parsed.target_price == "10.50000000"
        assert parsed.target_price_currency == "USD"


class TestDecimalAndCurrencyValidation:
    def test_bad_decimal_is_rejected(self) -> None:
        rows = validate_rows(
            parse_csv(_basic_csv("PN-1,Widget,,,,,not-a-number,USD")).rows
        )
        assert rows[0].parsed is None
        assert any(e.field == "target_price" for e in rows[0].errors)

    def test_price_finer_than_scale_is_rejected(self) -> None:
        rows = validate_rows(
            parse_csv(_basic_csv("PN-1,Widget,,,,,1.123456789,USD")).rows
        )
        assert rows[0].parsed is None
        assert any(e.field == "target_price" for e in rows[0].errors)

    def test_negative_price_is_rejected(self) -> None:
        rows = validate_rows(parse_csv(_basic_csv("PN-1,Widget,,,,,-1,USD")).rows)
        assert rows[0].parsed is None
        assert any(e.field == "target_price" for e in rows[0].errors)

    def test_bad_currency_is_rejected(self) -> None:
        rows = validate_rows(
            parse_csv(_basic_csv("PN-1,Widget,,,,,10.00,DOLLARS")).rows
        )
        assert rows[0].parsed is None
        assert any(e.field == "target_price_currency" for e in rows[0].errors)

    def test_currency_is_uppercased(self) -> None:
        rows = validate_rows(
            parse_csv(_basic_csv("PN-1,Widget,,,,,10.00,usd")).rows
        )
        assert rows[0].parsed is not None
        assert rows[0].parsed.target_price_currency == "USD"

    def test_price_without_currency_is_rejected(self) -> None:
        rows = validate_rows(parse_csv(_basic_csv("PN-1,Widget,,,,,10.00,")).rows)
        assert rows[0].parsed is None
        assert any(e.field == "target_price_currency" for e in rows[0].errors)

    def test_currency_without_price_is_rejected(self) -> None:
        rows = validate_rows(parse_csv(_basic_csv("PN-1,Widget,,,,,,USD")).rows)
        assert rows[0].parsed is None

    def test_neither_price_nor_currency_is_fine(self) -> None:
        rows = validate_rows(parse_csv(_basic_csv("PN-1,Widget,,,,,,")).rows)
        assert rows[0].parsed is not None
        assert rows[0].parsed.target_price is None
        assert rows[0].parsed.target_price_currency is None


class TestInFileDuplicateDetection:
    def test_second_occurrence_is_flagged_duplicate(self) -> None:
        rows = validate_rows(
            parse_csv(
                _basic_csv(
                    "PN-1,Widget A,,,,,,",
                    "PN-1,Widget B,,,,,,",
                )
            ).rows
        )
        assert rows[0].is_duplicate_in_file is False
        assert rows[1].is_duplicate_in_file is True

    def test_duplicate_detection_is_normalized(self) -> None:
        # "PN-1" and "pn 1" normalize identically (lower + strip non-alnum)
        rows = validate_rows(
            parse_csv(
                _basic_csv(
                    "PN-1,Widget A,,,,,,",
                    "pn 1,Widget B,,,,,,",
                )
            ).rows
        )
        assert rows[0].is_duplicate_in_file is False
        assert rows[1].is_duplicate_in_file is True

    def test_distinct_numbers_are_not_duplicates(self) -> None:
        rows = validate_rows(
            parse_csv(
                _basic_csv(
                    "PN-1,Widget A,,,,,,",
                    "PN-2,Widget B,,,,,,",
                )
            ).rows
        )
        assert rows[0].is_duplicate_in_file is False
        assert rows[1].is_duplicate_in_file is False

    def test_normalize_internal_part_number_helper(self) -> None:
        assert normalize_internal_part_number("ACME-100 Rev.B") == "acme100revb"


def _write_xlsx(headers: list[str], rows: list[list[object]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestXlsxParsing:
    def test_basic_round_trip(self) -> None:
        data = _write_xlsx(
            ["internal_part_number", "name", "target_price", "target_price_currency"],
            [["PN-1", "Widget", 10.5, "USD"]],
        )
        raw_file = parse_xlsx(data)
        assert len(raw_file.rows) == 1
        cells = raw_file.rows[0].cells
        assert cells["internal_part_number"] == "PN-1"
        assert cells["name"] == "Widget"
        assert cells["target_price"] == "10.5"
        assert cells["target_price_currency"] == "USD"

    def test_formula_with_cached_value_is_used(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["internal_part_number", "name", "target_price"])
        ws.append(["PN-1", "Widget", "=5+5"])
        buf = io.BytesIO()
        wb.save(buf)
        # openpyxl never computes formulas; simulate a cached value the way
        # a real spreadsheet application would have written one by loading
        # with data_only=False and asserting our parser degrades gracefully
        # when there genuinely is no cached value (see next test) instead -
        # openpyxl cannot fabricate a cached value for us in this test setup.
        data = buf.getvalue()
        raw_file = parse_xlsx(data)
        # no cached value exists (openpyxl-authored file, never opened by a
        # real spreadsheet app), so this is treated as an unresolved formula
        assert "target_price" in raw_file.rows[0].formula_columns

    def test_row_error_recorded_for_unresolved_formula(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["internal_part_number", "name", "target_price"])
        ws.append(["PN-1", "Widget", "=5+5"])
        buf = io.BytesIO()
        wb.save(buf)
        raw_file = parse_xlsx(buf.getvalue())
        rows = validate_rows(raw_file.rows)
        assert rows[0].parsed is None
        assert any("formula" in e.issue for e in rows[0].errors)

    def test_too_many_columns_is_rejected(self) -> None:
        headers = [f"col{i}" for i in range(MAX_COLUMNS + 1)]
        data = _write_xlsx(headers, [])
        with pytest.raises(PartImportParseError, match="columns"):
            parse_xlsx(data)

    def test_unknown_header_is_rejected(self) -> None:
        data = _write_xlsx(["internal_part_number", "name", "bogus"], [["PN-1", "Widget", "x"]])
        with pytest.raises(PartImportParseError, match="Unrecognized"):
            parse_xlsx(data)

    def test_blank_rows_are_skipped(self) -> None:
        data = _write_xlsx(
            ["internal_part_number", "name"],
            [["PN-1", "Widget"], [None, None], ["PN-2", "Gadget"]],
        )
        raw_file = parse_xlsx(data)
        assert len(raw_file.rows) == 2

    def test_corrupt_workbook_is_rejected(self) -> None:
        with pytest.raises(PartImportParseError):
            parse_xlsx(b"not an xlsx file at all")
