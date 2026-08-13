"""XLSX renderer: `ReportDocument` -> bytes, via `openpyxl` (docs/SPEC.md
§Reports and exports).

One worksheet, same header-block + block-by-block layout as
`csv_renderer.py` (kept structurally identical so a user comparing the two
formats of the same report sees the same content in the same order).
Header rows (the report title, each block's own heading, and each table's
column row) are bold for readability; this is presentation only; content
rules are identical to the CSV path.

Every cell passes through `app.reports.escape.escape_formula_cell` before
being written - see that module's docstring for why this matters even more
here than for CSV: `openpyxl` itself auto-detects a leading `=` on a string
`Cell.value` and stores it as a live formula unless the leading character
has already been neutralized.
"""

from __future__ import annotations

import io
from typing import Any

# no bundled type stubs, and pyproject.toml's mypy overrides are off-limits
# for this change - same accepted precedent app/importing/part_import_parser.py
# already documents for its own openpyxl import.
import openpyxl  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]

from app.reports.document import KeyValueBlock, ReportDocument, TableBlock, TextBlock
from app.reports.escape import escape_formula_cell

_BOLD = Font(bold=True)


def _write_row(ws: Any, cells: list[str], *, bold: bool = False) -> None:
    ws.append([escape_formula_cell(c) for c in cells])
    if bold:
        row_idx = ws.max_row
        for col_idx in range(1, len(cells) + 1):
            ws.cell(row=row_idx, column=col_idx).font = _BOLD


def render_xlsx(document: ReportDocument) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    _write_row(ws, ["Report", document.title], bold=True)
    _write_row(ws, ["Generated At", document.generated_at])
    _write_row(ws, ["Calculation Version", document.calculation_version])
    _write_row(ws, ["Methodology", document.methodology])
    _write_row(ws, ["Disclaimer", document.disclaimer])
    if document.missing_data:
        for item in document.missing_data:
            _write_row(ws, ["Missing Data", item])
    else:
        _write_row(ws, ["Missing Data", "None"])
    ws.append([])

    for block in document.blocks:
        _write_row(ws, [block.heading], bold=True)
        if isinstance(block, TableBlock):
            _write_row(ws, list(block.columns), bold=True)
            for data_row in block.rows:
                _write_row(ws, list(data_row))
        elif isinstance(block, KeyValueBlock):
            for key, value in block.pairs:
                _write_row(ws, [key, value])
        elif isinstance(block, TextBlock):
            for paragraph in block.paragraphs:
                _write_row(ws, [paragraph])
        ws.append([])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["render_xlsx"]
