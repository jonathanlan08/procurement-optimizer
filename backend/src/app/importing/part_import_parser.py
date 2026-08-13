"""CSV/XLSX parsing and row-level validation for part imports (SPEC §3;
docs/planning/03-api-contract.md §4.5 `POST /part-imports`).

Pure - no database session anywhere in this module. `app.services.
part_import_service.PartImportService` does the DB-dependent second pass
(existing-part duplicate detection, `unit_code` → `unit_definition_id`
resolution, persistence). Keeping this layer pure makes header validation,
decimal/currency parsing, formula-injection rejection, and in-file duplicate
detection unit-testable without a database
(backend/tests/unit/test_part_import_parser.py).

Canonical import columns (the spec; the API contract does not
define its own column list, so the brief's list is authoritative here,
reconciled against `app.models.parts.Part`'s own nullability):
`internal_part_number` and `name` are required (matching `Part.
internal_part_number`/`Part.name`, both `NOT NULL`); `manufacturer_part_
number`, `description`, `category`, `unit_code`, `target_price`, `target_
price_currency` are optional. `target_price`/`target_price_currency` must
be both-or-neither, mirroring the DB CHECK `ck_parts_target_price_currency_
paired` (migration 0004) via the identical rule already used by `app.
schemas.parts.PartCreate`.

`unit_code` is listed as optional even though `Part.unit_definition_id` is
`NOT NULL`: a blank/absent `unit_code` is not rejected here - the service
resolves it to the standard catalogue's `"each"` unit (`app.seed.units_
catalog.STANDARD_UNIT_CATALOG`'s canonical count-dimension entry, factor 1
by definition) as the sensible default for "no unit of measure specified",
and only *that* resolution failing (e.g. an org whose catalogue was never
seeded) makes the row invalid. See PartImportService._resolve_unit_code.

Security (the spec, non-negotiable):
- **CSV formula injection.** A cell whose first character is one of
  ``=+-@`` is rejected as a row-level error *unless* the entire cell is a
  plain signed number (``-10.5``, ``+5`` are legitimate negative/positive
  numbers; ``=cmd|'/c calc'!A1``, ``-2+3+cmd|' /C calc'!A1`` are not).
  Applied to every cell of every row, for both CSV and XLSX - a string cell
  in an XLSX sheet can carry the identical payload if the sheet is ever
  round-tripped back through a spreadsheet application, so the same check
  guards both formats rather than only the bullet that names CSV
  explicitly.
- **Encoding.** UTF-8 first, BOM-tolerant (`"utf-8-sig"` strips a leading
  BOM if present and is a strict superset of plain UTF-8 decoding).
  Latin-1 fallback on a `UnicodeDecodeError` - Latin-1 can decode any byte
  sequence, so it never itself raises; `RawFile.encoding_fallback` reports
  when it was used so the caller can surface it, never silently.
- **Row/column caps.** 10,000 data rows / 50 columns. Exceeding either
  rejects the *whole file* (`PartImportParseError`) rather than silently
  truncating - this far outside normal usage reads as abuse, not a file to
  quietly accept part of.
- **XLSX.** Loaded with `openpyxl(read_only=True, data_only=True)` so this
  process never evaluates a formula - only cached values are read. A
  formula cell with no cached value is indistinguishable from a genuinely
  blank cell in a single `read_only+data_only` pass, so `parse_xlsx` loads
  the workbook a *second* time with `data_only=False` (still `read_only`,
  still not evaluating anything - this second pass only reads which cells
  store an `=`-prefixed formula string) purely to detect "formula, no
  cached value", which is reported as a row error by design.
  Defined names and macros are not detectable via openpyxl (the
  brief says so explicitly) and `.xlsm` is outside the accepted extensions
  (`app.api.v1.part_imports._EXTENSION_FORMAT`), so macros never reach this
  parser at all.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal

# no bundled type stubs, and pyproject.toml's mypy overrides are off-limits for this change
import openpyxl  # type: ignore[import-untyped]

from app.core.money import UNIT_PRICE_SCALE, MoneyError, parse_at_scale

MAX_ROWS = 10_000
MAX_COLUMNS = 50

REQUIRED_COLUMNS: frozenset[str] = frozenset({"internal_part_number", "name"})
OPTIONAL_COLUMNS: frozenset[str] = frozenset(
    {
        "manufacturer_part_number",
        "description",
        "category",
        "unit_code",
        "target_price",
        "target_price_currency",
    }
)
CANONICAL_COLUMNS: frozenset[str] = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

_CURRENCY_RE = re.compile(r"^[A-Z]{3}$")
_INJECTION_LEAD = frozenset("=+-@")
# a signed plain number: "-10.5", "+5", "5", ".5" - anything else after a
# leading =+-@ is treated as a formula-injection attempt, not a number
_SAFE_SIGNED_NUMBER_RE = re.compile(r"^[+-]?(\d+\.?\d*|\.\d+)$")
_NON_ALNUM_RE = re.compile(r"[^a-z0-9]")


class PartImportParseError(Exception):
    """File-level parse failure. Nothing is persisted when this is raised -
    header validation, encoding, size/shape caps, and corrupt-workbook
    failures all abort before any batch or row is created."""


def normalize_internal_part_number(raw: str) -> str:
    """Mirrors `app.repositories.part_repository._normalize` / `Part.
    normalized_key` (lower + strip non-alphanumerics) so in-file duplicate
    detection agrees with the DB's own case/punctuation-insensitive
    uniqueness rule (`uq_parts_org_ipn_active`, migration 0004)."""
    return _NON_ALNUM_RE.sub("", raw.lower())


def _is_formula_injection(value: str) -> bool:
    s = value.strip()
    if not s or s[0] not in _INJECTION_LEAD:
        return False
    return _SAFE_SIGNED_NUMBER_RE.match(s) is None


@dataclass(frozen=True, slots=True)
class RawRow:
    """One data row as decoded strings, keyed by canonical column name.
    `formula_columns` is XLSX-only: columns whose cell stores a formula with
    no cached value (see module docstring)."""

    row_number: int  # 1-based; counts data rows only, header excluded
    cells: dict[str, str]
    formula_columns: frozenset[str] = field(default_factory=frozenset)


@dataclass(frozen=True, slots=True)
class RawFile:
    rows: list[RawRow]
    encoding_fallback: bool = False


@dataclass(frozen=True, slots=True)
class RowError:
    field: str | None
    issue: str

    def to_dict(self) -> dict[str, str | None]:
        return {"field": self.field, "issue": self.issue}


@dataclass(frozen=True, slots=True)
class ParsedFields:
    """Field values after per-row validation, before DB-dependent resolution
    (unit_code -> unit_definition_id, existing-part duplicate check -
    app.services.part_import_service). `target_price` is a decimal string at
    UNIT_PRICE_SCALE (wire-ready, `app.core.money.to_wire` shape), never a
    float, matching the codebase's money policy."""

    internal_part_number: str
    manufacturer_part_number: str | None
    name: str
    description: str | None
    category: str | None
    unit_code: str | None
    target_price: str | None
    target_price_currency: str | None

    def to_json(self) -> dict[str, str | None]:
        return {
            "internal_part_number": self.internal_part_number,
            "manufacturer_part_number": self.manufacturer_part_number,
            "name": self.name,
            "description": self.description,
            "category": self.category,
            "unit_code": self.unit_code,
            "target_price": self.target_price,
            "target_price_currency": self.target_price_currency,
        }


@dataclass(frozen=True, slots=True)
class ValidatedRow:
    row_number: int
    raw: dict[str, str]
    parsed: ParsedFields | None  # None when `errors` is non-empty
    errors: list[RowError]
    normalized_internal_part_number: str | None
    is_duplicate_in_file: bool


def _decode_csv_bytes(data: bytes) -> tuple[str, bool]:
    """Returns (text, fallback_used). UTF-8 (BOM-tolerant) first, Latin-1
    fallback - see module docstring."""
    try:
        return data.decode("utf-8-sig"), False
    except UnicodeDecodeError:
        return data.decode("latin-1"), True


def _validate_headers(headers: list[str]) -> dict[int, str]:
    """Column index -> canonical field name. Raises `PartImportParseError`
    on an unrecognized, duplicated, or missing-required header - SPEC §3's
    "header validation" is a whole-file gate, not a per-row concern."""
    seen: dict[str, int] = {}
    mapping: dict[int, str] = {}
    unknown: list[str] = []
    for idx, raw_header in enumerate(headers):
        name = raw_header.strip().lower()
        if not name:
            continue  # blank trailing header cell (common with trailing commas)
        if name not in CANONICAL_COLUMNS:
            unknown.append(raw_header)
            continue
        if name in seen:
            raise PartImportParseError(f"Duplicate column header {raw_header!r}.")
        seen[name] = idx
        mapping[idx] = name
    if unknown:
        raise PartImportParseError(
            "Unrecognized column header(s): "
            + ", ".join(repr(h) for h in unknown)
            + ". Allowed columns: "
            + ", ".join(sorted(CANONICAL_COLUMNS))
            + "."
        )
    missing = REQUIRED_COLUMNS - seen.keys()
    if missing:
        raise PartImportParseError(
            f"Missing required column header(s): {', '.join(sorted(missing))}."
        )
    return mapping


def parse_csv(data: bytes) -> RawFile:
    text_data, fallback = _decode_csv_bytes(data)
    reader = csv.reader(io.StringIO(text_data))
    try:
        header_row = next(reader)
    except StopIteration:
        raise PartImportParseError("File is empty; no header row found.") from None

    if len(header_row) > MAX_COLUMNS:
        raise PartImportParseError(
            f"File has {len(header_row)} columns, exceeding the maximum of {MAX_COLUMNS}."
        )
    column_map = _validate_headers(header_row)

    rows: list[RawRow] = []
    for row_number, record in enumerate(reader, start=1):
        if row_number > MAX_ROWS:
            raise PartImportParseError(
                f"File has more than {MAX_ROWS} data rows, exceeding the maximum allowed."
            )
        if not record or all(not cell.strip() for cell in record):
            continue  # skip fully blank lines (common trailing-EOF artifact)
        cells: dict[str, str] = {}
        for idx, field_name in column_map.items():
            raw_value = record[idx].strip() if idx < len(record) else ""
            cells[field_name] = raw_value
        rows.append(RawRow(row_number=row_number, cells=cells))
    return RawFile(rows=rows, encoding_fallback=fallback)


def _xlsx_cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        # via str() first to avoid binary-float artifacts (Decimal(0.1) != "0.1")
        return format(Decimal(str(value)), "f")
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, datetime | date):
        return value.isoformat()
    return str(value).strip()


def parse_xlsx(data: bytes) -> RawFile:
    try:
        wb_values = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises assorted exception types for corrupt input
        raise PartImportParseError(f"Could not open XLSX file: {exc}") from exc

    try:
        ws_values = wb_values.active
        if ws_values is None:
            raise PartImportParseError("XLSX file has no active worksheet.")

        try:
            wb_formulas = openpyxl.load_workbook(
                io.BytesIO(data), read_only=True, data_only=False
            )
        except Exception as exc:
            raise PartImportParseError(f"Could not open XLSX file: {exc}") from exc

        try:
            ws_formulas = wb_formulas[ws_values.title]

            value_rows = ws_values.iter_rows(values_only=True)
            formula_rows = ws_formulas.iter_rows(values_only=True)

            try:
                header_row = next(value_rows)
                next(formula_rows)
            except StopIteration:
                raise PartImportParseError("File is empty; no header row found.") from None

            header_cells = ["" if v is None else str(v) for v in header_row]
            if len(header_cells) > MAX_COLUMNS:
                raise PartImportParseError(
                    f"File has {len(header_cells)} columns, exceeding the maximum of "
                    f"{MAX_COLUMNS}."
                )
            column_map = _validate_headers(header_cells)

            rows: list[RawRow] = []
            row_number = 0
            for value_record, formula_record in zip(value_rows, formula_rows, strict=False):
                row_number += 1
                if row_number > MAX_ROWS:
                    raise PartImportParseError(
                        f"File has more than {MAX_ROWS} data rows, exceeding the maximum "
                        "allowed."
                    )
                if value_record is None or all(v is None for v in value_record):
                    continue

                cells: dict[str, str] = {}
                formula_columns: set[str] = set()
                for idx, field_name in column_map.items():
                    raw_value = value_record[idx] if idx < len(value_record) else None
                    formula_value = (
                        formula_record[idx]
                        if formula_record is not None and idx < len(formula_record)
                        else None
                    )
                    is_unresolved_formula = (
                        isinstance(formula_value, str)
                        and formula_value.startswith("=")
                        and raw_value is None
                    )
                    if is_unresolved_formula:
                        formula_columns.add(field_name)
                        cells[field_name] = ""
                        continue
                    cells[field_name] = _xlsx_cell_to_str(raw_value)
                rows.append(
                    RawRow(
                        row_number=row_number,
                        cells=cells,
                        formula_columns=frozenset(formula_columns),
                    )
                )
            return RawFile(rows=rows, encoding_fallback=False)
        finally:
            wb_formulas.close()
    finally:
        wb_values.close()


def _validate_row(raw_row: RawRow) -> tuple[ParsedFields | None, list[RowError], str | None]:
    errors: list[RowError] = []

    for field_name, value in raw_row.cells.items():
        if value and _is_formula_injection(value):
            errors.append(
                RowError(
                    field=field_name,
                    issue=(
                        "value begins with a spreadsheet formula character (=, +, -, @) "
                        "and is not a plain number; rejected as a formula-injection risk"
                    ),
                )
            )
    for field_name in sorted(raw_row.formula_columns):
        errors.append(
            RowError(
                field=field_name,
                issue="cell contains a formula with no cached value; formulas are never evaluated",
            )
        )

    internal_part_number = raw_row.cells.get("internal_part_number", "").strip()
    name = raw_row.cells.get("name", "").strip()
    if not internal_part_number:
        errors.append(RowError(field="internal_part_number", issue="required value is missing"))
    if not name:
        errors.append(RowError(field="name", issue="required value is missing"))

    manufacturer_part_number = raw_row.cells.get("manufacturer_part_number", "").strip() or None
    description = raw_row.cells.get("description", "").strip() or None
    category = raw_row.cells.get("category", "").strip() or None
    unit_code_raw = raw_row.cells.get("unit_code", "").strip() or None

    target_price_raw = raw_row.cells.get("target_price", "").strip() or None
    target_price_currency_raw = raw_row.cells.get("target_price_currency", "").strip() or None

    target_price: str | None = None
    price_field_error = False
    if target_price_raw is not None:
        try:
            parsed_price = parse_at_scale(target_price_raw, UNIT_PRICE_SCALE)
        except MoneyError as exc:
            errors.append(RowError(field="target_price", issue=str(exc)))
            price_field_error = True
        else:
            if parsed_price < 0:
                errors.append(RowError(field="target_price", issue="must not be negative"))
                price_field_error = True
            else:
                target_price = format(parsed_price, "f")

    target_price_currency: str | None = None
    currency_field_error = False
    if target_price_currency_raw is not None:
        normalized_currency = target_price_currency_raw.upper()
        if not _CURRENCY_RE.match(normalized_currency):
            errors.append(
                RowError(field="target_price_currency", issue="must be a 3-letter ISO 4217 code")
            )
            currency_field_error = True
        else:
            target_price_currency = normalized_currency

    if (
        (target_price is None) != (target_price_currency is None)
        and not price_field_error
        and not currency_field_error
    ):
        errors.append(
            RowError(
                field="target_price_currency",
                issue="target_price and target_price_currency must both be set or both be absent",
            )
        )
        target_price = None
        target_price_currency = None

    normalized_ipn = (
        normalize_internal_part_number(internal_part_number) if internal_part_number else None
    )

    if errors:
        return None, errors, normalized_ipn

    parsed = ParsedFields(
        internal_part_number=internal_part_number,
        manufacturer_part_number=manufacturer_part_number,
        name=name,
        description=description,
        category=category,
        unit_code=unit_code_raw,
        target_price=target_price,
        target_price_currency=target_price_currency,
    )
    return parsed, errors, normalized_ipn


def validate_rows(raw_rows: list[RawRow]) -> list[ValidatedRow]:
    """Per-row field validation plus in-file duplicate detection by
    normalized `internal_part_number` (SPEC §3). The *first* row bearing a
    given normalized number is not a duplicate; every later row with the
    same normalized number is, independent of either row's own validity."""
    seen_normalized: set[str] = set()
    results: list[ValidatedRow] = []
    for raw_row in raw_rows:
        parsed, errors, normalized_ipn = _validate_row(raw_row)
        is_duplicate_in_file = False
        if normalized_ipn is not None:
            if normalized_ipn in seen_normalized:
                is_duplicate_in_file = True
            else:
                seen_normalized.add(normalized_ipn)
        results.append(
            ValidatedRow(
                row_number=raw_row.row_number,
                raw=dict(raw_row.cells),
                parsed=parsed,
                errors=errors,
                normalized_internal_part_number=normalized_ipn,
                is_duplicate_in_file=is_duplicate_in_file,
            )
        )
    return results


__all__ = [
    "CANONICAL_COLUMNS",
    "MAX_COLUMNS",
    "MAX_ROWS",
    "OPTIONAL_COLUMNS",
    "REQUIRED_COLUMNS",
    "ParsedFields",
    "PartImportParseError",
    "RawFile",
    "RawRow",
    "RowError",
    "ValidatedRow",
    "normalize_internal_part_number",
    "parse_csv",
    "parse_xlsx",
    "validate_rows",
]
