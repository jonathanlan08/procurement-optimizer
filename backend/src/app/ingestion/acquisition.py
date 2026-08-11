"""Stage 4 — text/table acquisition (docs/planning/04-document-pipeline.md §5).

Pure function over already content-validated bytes (`app.ingestion.file_validation` has
already sniffed and admitted `kind` from magic bytes before this module ever runs): parses
the buffer in memory and returns one `PageText` per page/sheet. No I/O beyond parsing, no
network access, no OCR execution, no database access, nothing written to disk.

This module deliberately does NOT interpret, filter, redact, or judge the text it
acquires — trust-boundary framing (the nonce envelope) and injection-canary scanning both
happen downstream, in `app.providers.extraction.envelope`, over the text this module
returns. A document that contains an injection string must see that string survive
acquisition byte-for-byte; flagging it is the service layer's job, not this one's
(04-document-pipeline.md §6 point 6: "filtering the text would be worse").

OCR is provider-abstracted (04-document-pipeline.md §5's `OcrProvider` row) and, per SPEC
"External-service strategy", v0.1.0 ships no bundled OCR engine — wiring a real OCR SDK
(Tesseract, a cloud OCR API, ...) is out of scope for this portfolio build. PNG/JPEG
documents therefore acquire as a single page with empty text and `used_ocr=True`; a real
OCR adapter would honor the same `used_ocr=True` contract and simply fill in `text` with
its recognized characters. The mock extraction provider
(`app.providers.extraction.mock.MockExtractionProvider`) knows to answer from its fixture
registry rather than try to read pixels when it sees this flag.

Table extraction for PDFs is two-layered: `pypdf` supplies the primary per-page text (fast,
always available, works even when a page has no ruled table); `pdfplumber` additionally
locates ruled/gridded table regions and appends each one to that page's text as
pipe-separated rows, so structured line-item tables survive even when `pypdf`'s
flowing-text extraction loses column alignment. `pdfplumber` is best-effort: a failure
there (e.g. a structurally unusual PDF `pdfplumber` chokes on) must not discard the `pypdf`
text that already succeeded.
"""

from __future__ import annotations

import contextlib
import csv
import io
from dataclasses import dataclass

import openpyxl  # type: ignore[import-untyped]
import pdfplumber
from pypdf import PdfReader

from app.ingestion.file_validation import DocumentKind

# Acquisition-layer caps. Distinct from, and smaller than, Stage-2 content-validation caps
# admitted at upload time (04-document-pipeline.md §3: CSV 50k rows / 200 cols) — those
# gate what is allowed into storage at all; these bound the cost of *this* parse. The
# numbers mirror `app.importing.part_import_parser`'s identical row/column caps for the
# same class of tabular acquisition, reused here for both XLSX (per the delegating task)
# and CSV (same row/column shape, same cost concern).
MAX_ACQUIRED_ROWS = 10_000
MAX_ACQUIRED_COLUMNS = 50


class AcquisitionLimitError(ValueError):
    """A document exceeded this module's row/column acquisition cap. Message is safe for
    display; never echoes raw document content."""


@dataclass(frozen=True, slots=True)
class PageText:
    """One page's (or sheet's) acquired text — exactly what the extraction provider will
    see for this page. `used_ocr=True` marks pages whose `text` was not read from a native
    text layer (see module docstring)."""

    page_number: int
    text: str
    used_ocr: bool = False


def acquire_pages(data: bytes, kind: DocumentKind) -> list[PageText]:
    """Dispatch to the acquisition strategy for `kind`. `kind` is trusted input — already
    sniffed from magic bytes by `file_validation.sniff_kind` — and is never re-derived
    from `data` here."""
    if kind is DocumentKind.PDF:
        return _acquire_pdf(data)
    if kind in (DocumentKind.PNG, DocumentKind.JPEG):
        return _acquire_image()
    if kind is DocumentKind.CSV:
        return _acquire_csv(data)
    if kind is DocumentKind.XLSX:
        return _acquire_xlsx(data)
    raise AssertionError(f"unhandled DocumentKind: {kind!r}")  # exhaustive over the StrEnum


def _acquire_pdf(data: bytes) -> list[PageText]:
    reader = PdfReader(io.BytesIO(data))
    texts: list[str] = [(page.extract_text() or "") for page in reader.pages]

    # Table extraction is a best-effort supplement to the pypdf text above; a pdfplumber
    # failure on a structurally-odd PDF must not discard text pypdf already produced.
    with contextlib.suppress(Exception), pdfplumber.open(io.BytesIO(data)) as pdf:
        for index, plumber_page in enumerate(pdf.pages):
            if index >= len(texts):
                break
            for table in plumber_page.extract_tables():
                table_text = _table_to_pipe_rows(table)
                if not table_text:
                    continue
                texts[index] = f"{texts[index]}\n{table_text}" if texts[index] else table_text

    return [PageText(page_number=i, text=t) for i, t in enumerate(texts, start=1)]


def _table_to_pipe_rows(table: list[list[str | None]]) -> str:
    rows = ["|".join(cell if cell is not None else "" for cell in row) for row in table]
    return "\n".join(rows)


def _acquire_image() -> list[PageText]:
    return [PageText(page_number=1, text="", used_ocr=True)]


def _decode_text(data: bytes) -> str:
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        return data.decode("latin-1")  # never raises: Latin-1 maps every byte value


def _acquire_csv(data: bytes) -> list[PageText]:
    text = _decode_text(data)
    for row_count, row in enumerate(csv.reader(io.StringIO(text)), start=1):
        if row_count > MAX_ACQUIRED_ROWS:
            raise AcquisitionLimitError(f"CSV exceeds the {MAX_ACQUIRED_ROWS}-row acquisition cap")
        if len(row) > MAX_ACQUIRED_COLUMNS:
            raise AcquisitionLimitError(
                f"CSV exceeds the {MAX_ACQUIRED_COLUMNS}-column acquisition cap"
            )
    return [PageText(page_number=1, text=text)]


def _acquire_xlsx(data: bytes) -> list[PageText]:
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        pages: list[PageText] = []
        for page_number, sheet_name in enumerate(workbook.sheetnames, start=1):
            worksheet = workbook[sheet_name]
            lines: list[str] = []
            for row_index, row in enumerate(worksheet.iter_rows(), start=1):
                if row_index > MAX_ACQUIRED_ROWS:
                    raise AcquisitionLimitError(
                        f"sheet {sheet_name!r} exceeds the {MAX_ACQUIRED_ROWS}-row "
                        "acquisition cap"
                    )
                if len(row) > MAX_ACQUIRED_COLUMNS:
                    raise AcquisitionLimitError(
                        f"sheet {sheet_name!r} exceeds the {MAX_ACQUIRED_COLUMNS}-column "
                        "acquisition cap"
                    )
                cells = [_cell_text(cell.value) for cell in row]
                lines.append("|".join(cells))
            pages.append(PageText(page_number=page_number, text="\n".join(lines)))
        return pages
    finally:
        workbook.close()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


__all__ = [
    "MAX_ACQUIRED_COLUMNS",
    "MAX_ACQUIRED_ROWS",
    "AcquisitionLimitError",
    "PageText",
    "acquire_pages",
]
