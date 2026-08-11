"""Report API integration tests (docs/planning/03-api-contract.md §4.18,
app/services/report_service.py, docs/planning/02-erd.md GENERATED_REPORTS
box + §11).

Fixture pattern mirrors test_briefs_api.py: a committed org + one user per
role, built directly against the migrated database, driven through the HTTP
API. `_setup_two_supplier_case`/`_create_and_complete_scenario` are copied
from that file's own signature case (Acme Low-Price vs. Beta Premium, 500
units) rather than imported, the same "duplicate the small fixture helper
set per test module" precedent `test_documents_api.py`/`test_briefs_api.py`
already establish for each other. The `client` fixture additionally wires a
real (filesystem, tmp-dir-scoped) storage provider, the same technique
`test_documents_api.py`'s own `client` fixture uses, since report content
round-trips through `app.providers.storage`.
"""

from __future__ import annotations

import hashlib
import io
import uuid
from collections.abc import Generator
from datetime import UTC, datetime
from decimal import Decimal
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import Engine
from sqlalchemy import text as sqltext
from sqlalchemy.orm import Session as SaSession

from app.core.config import Environment, Settings
from app.core.security import hash_password
from app.main import create_app
from app.models.identity import Role

ORIGIN = "http://localhost:5173"
PASSWORD = "correct-horse-battery-report"

ASSUMPTIONS: dict[str, Any] = {
    "quality_risk_rate": "0.02",
    "delay_risk_per_day": "0",
    "promised_lead_time_days": "0",
    "required_lead_time_days": "0",
    "annual_rate": "0.08",
    "baseline_terms_days": "30",
    "tariff_rate": "0.30",
    "assume_missing_costs_zero": True,
}


@pytest.fixture(scope="module")
def client(
    database_url: str, migrated_engine: Engine, tmp_path_factory: pytest.TempPathFactory
) -> Generator[TestClient, None, None]:
    storage_root = tmp_path_factory.mktemp("report-storage")
    settings = Settings(
        environment=Environment.TEST,
        database_url=database_url,
        allowed_origins=[ORIGIN],
        rate_limit_auth_per_minute=100_000,
        rate_limit_per_minute=100_000,
        storage_root=str(storage_root),
    )
    app = create_app(settings)
    with TestClient(app, base_url="http://testserver") as c:
        yield c
    app.state.engine.dispose()


def _make_org_with_members(migrated_engine: Engine, roles: list[Role]) -> dict[str, Any]:
    from app.models.identity import Organization, OrganizationMembership, User

    suffix = uuid.uuid4().hex[:10]
    now = datetime.now(UTC)
    with SaSession(migrated_engine) as s:
        org = Organization(
            id=uuid.uuid4(),
            slug=f"report-org-{suffix}",
            name=f"Report Org {suffix}",
            base_currency="USD",
            is_demo=False,
            created_at=now,
            updated_at=now,
        )
        s.add(org)
        users: dict[str, dict[str, str]] = {}
        for role in roles:
            email = f"{role.value}-{suffix}-{uuid.uuid4().hex[:6]}@example.com"
            user = User(
                id=uuid.uuid4(),
                email=email,
                password_hash=hash_password(PASSWORD),
                full_name=f"{role.value.title()} User",
                created_at=now,
                updated_at=now,
            )
            s.add(user)
            s.add(
                OrganizationMembership(
                    id=uuid.uuid4(),
                    organization_id=org.id,
                    user_id=user.id,
                    role=role,
                    accepted_at=now,
                    created_at=now,
                    updated_at=now,
                )
            )
            users[role.value] = {"email": email, "user_id": str(user.id)}
        s.commit()
        return {"org_id": str(org.id), "users": users}


@pytest.fixture()
def org_a(migrated_engine: Engine) -> dict[str, Any]:
    return _make_org_with_members(
        migrated_engine, [Role.ANALYST, Role.ADMINISTRATOR, Role.VIEWER]
    )


@pytest.fixture()
def org_b(migrated_engine: Engine) -> dict[str, Any]:
    return _make_org_with_members(
        migrated_engine, [Role.ANALYST, Role.ADMINISTRATOR, Role.VIEWER]
    )


def _login_as(client: TestClient, org: dict[str, Any], role: Role) -> dict[str, str]:
    creds = org["users"][role.value]
    resp = client.post(
        "/api/v1/auth/login",
        json={"email": creds["email"], "password": PASSWORD},
        headers={"Origin": ORIGIN},
    )
    assert resp.status_code == 200, resp.text
    return resp.json()


def _headers(login_body: dict[str, str]) -> dict[str, str]:
    return {"Origin": ORIGIN, "X-CSRF-Token": login_body["csrf_token"]}


def _seed_unit(migrated_engine: Engine, organization_id: str) -> str:
    from app.models.units import UnitDefinition, UnitDimension

    unit_id = uuid.uuid4()
    code = f"unit-{uuid.uuid4().hex[:12]}"
    with SaSession(migrated_engine) as s:
        s.add(
            UnitDefinition(
                id=unit_id,
                organization_id=uuid.UUID(organization_id),
                code=code,
                display_name=code,
                dimension=UnitDimension.COUNT,
                to_canonical_factor=Decimal("1"),
                is_user_defined=True,
            )
        )
        s.commit()
    return str(unit_id)


def _create_part(client: TestClient, headers: dict[str, str], unit_id: str) -> dict[str, Any]:
    resp = client.post(
        "/api/v1/parts",
        json={
            "internal_part_number": f"PN-{uuid.uuid4().hex[:8].upper()}",
            "name": "Report test bracket",
            "unit_definition_id": unit_id,
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()  # type: ignore[no-any-return]


def _create_supplier(client: TestClient, headers: dict[str, str], *, name: str) -> dict[str, Any]:
    resp = client.post(
        "/api/v1/suppliers",
        json={
            "code": f"SUP-{uuid.uuid4().hex[:8].upper()}",
            "name": name,
            "country_code": "US",
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()  # type: ignore[no-any-return]


def _create_rfq(
    client: TestClient, headers: dict[str, str], part_id: str, *, quantity: str
) -> dict[str, Any]:
    resp = client.post(
        "/api/v1/rfqs",
        json={
            "name": f"RFQ-{uuid.uuid4().hex[:8].upper()}",
            "internal_reference": f"REF-{uuid.uuid4().hex[:8].upper()}",
            "base_currency": "USD",
            "due_date": "2026-12-01",
            "lines": [{"part_id": part_id, "required_quantity": quantity}],
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()  # type: ignore[no-any-return]


def _invite_supplier(
    client: TestClient, headers: dict[str, str], rfq_id: str, supplier_id: str
) -> None:
    resp = client.post(
        f"/api/v1/rfqs/{rfq_id}/suppliers",
        json={"supplier_ids": [supplier_id]},
        headers=headers,
    )
    assert resp.status_code == 201, resp.text


def _set_rfq_status(
    client: TestClient, headers: dict[str, str], rfq_id: str, to_status: str
) -> None:
    resp = client.post(
        f"/api/v1/rfqs/{rfq_id}/status",
        json={"to_status": to_status, "reason": "test"},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text


def _create_quote(
    client: TestClient,
    headers: dict[str, str],
    rfq: dict[str, Any],
    supplier_id: str,
    unit_id: str,
    *,
    quantity: str,
    unit_price: str,
    shipping_cost: str = "0.000000",
) -> dict[str, Any]:
    line: dict[str, Any] = {
        "matched_rfq_line_id": rfq["lines"][0]["id"],
        "description": "Report test line",
        "quantity": quantity,
        "unit_definition_id": unit_id,
        "unit_price": unit_price,
        "shipping_cost": shipping_cost,
        "moq": "1.000000",
        "lead_time_days": 14,
        "country_of_origin": "US",
    }
    body: dict[str, Any] = {
        "supplier_id": supplier_id,
        "quote_date": "2026-08-01",
        "currency": "USD",
        "lines": [line],
        "terms": {"payment_terms": "Net 30", "payment_terms_days": 30},
    }
    resp = client.post(f"/api/v1/rfqs/{rfq['id']}/quotes", json=body, headers=headers)
    assert resp.status_code == 201, resp.text
    return resp.json()  # type: ignore[no-any-return]


def _setup_two_supplier_case(
    client: TestClient,
    headers: dict[str, str],
    migrated_engine: Engine,
    org: dict[str, Any],
    *,
    supplier_a_name: str = "Acme Low-Price",
) -> dict[str, Any]:
    """Acme Low-Price (unit_price 8.00, shipping 3000 -> high landed cost)
    vs. Beta Premium (unit_price 11.00, shipping 100 -> low landed cost),
    500 units — same signature case test_briefs_api.py/test_scenarios_api.py
    build, so Beta always ends up the lower-landed-cost alternative."""
    unit_id = _seed_unit(migrated_engine, org["org_id"])
    part = _create_part(client, headers, unit_id)
    rfq = _create_rfq(client, headers, part["id"], quantity="500.000000")
    supplier_a = _create_supplier(client, headers, name=supplier_a_name)
    supplier_b = _create_supplier(client, headers, name="Beta Premium")
    _invite_supplier(client, headers, rfq["id"], supplier_a["id"])
    _invite_supplier(client, headers, rfq["id"], supplier_b["id"])
    _set_rfq_status(client, headers, rfq["id"], "open")
    quote_a = _create_quote(
        client, headers, rfq, supplier_a["id"], unit_id,
        quantity="500.000000", unit_price="8.00000000", shipping_cost="3000.000000",
    )
    quote_b = _create_quote(
        client, headers, rfq, supplier_b["id"], unit_id,
        quantity="500.000000", unit_price="11.00000000", shipping_cost="100.000000",
    )
    return {
        "rfq": rfq, "supplier_a": supplier_a, "supplier_b": supplier_b,
        "quote_a": quote_a, "quote_b": quote_b,
    }


def _create_and_complete_scenario(
    client: TestClient,
    headers: dict[str, str],
    rfq_id: str,
    *,
    strategy: str = "lowest_landed_cost",
) -> dict[str, Any]:
    resp = client.post(
        f"/api/v1/rfqs/{rfq_id}/comparison-scenarios",
        json={
            "name": f"Scenario ({strategy})",
            "strategy": strategy,
            "assumptions": ASSUMPTIONS,
            "constraints": {},
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["state"] == "complete"
    return body  # type: ignore[no-any-return]


def _generate_brief(
    client: TestClient, headers: dict[str, str], scenario_id: str, supplier_id: str
) -> dict[str, Any]:
    resp = client.post(
        f"/api/v1/comparison-scenarios/{scenario_id}/negotiation-briefs",
        json={"supplier_id": supplier_id},
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()  # type: ignore[no-any-return]


def _generate_report(
    client: TestClient,
    headers: dict[str, str],
    *,
    scenario_id: str,
    report_type: str,
    format: str,
    parameters: dict[str, Any] | None = None,
) -> Any:
    body: dict[str, Any] = {
        "scenario_id": scenario_id,
        "report_type": report_type,
        "format": format,
    }
    if parameters is not None:
        body["parameters"] = parameters
    return client.post("/api/v1/reports", json=body, headers=headers)


# ---------------------------------------------------------------------------


class TestGenerateReportHappyPath:
    @pytest.mark.parametrize("format", ["csv", "xlsx"])
    def test_supplier_comparison(
        self,
        client: TestClient,
        org_a: dict[str, Any],
        migrated_engine: Engine,
        format: str,
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])

        resp = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="supplier_comparison", format=format,
        )
        assert resp.status_code == 201, resp.text
        body = resp.json()
        assert body["state"] == "ready"
        assert body["report_type"] == "supplier_comparison"
        assert body["format"] == format
        assert body["scenario_id"] == scenario["id"]
        assert body["size_bytes"] > 0
        assert body["content_sha256"] is not None
        assert len(body["content_sha256"]) == 64  # sha256 hex digest
        assert body["purged"] is False
        assert body["error_message"] is None
        assert body["calculation_version"] == scenario["calculation_version"]

    def test_cfo_recommendation_pdf(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])

        resp = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="cfo_recommendation", format="pdf",
        )
        assert resp.status_code == 201, resp.text
        body = resp.json()
        assert body["state"] == "ready"
        assert body["format"] == "pdf"

    def test_negotiation_brief_pdf_default_latest(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])
        _generate_brief(client, headers, scenario["id"], ctx["supplier_a"]["id"])

        resp = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="negotiation_brief", format="pdf",
        )
        assert resp.status_code == 201, resp.text
        assert resp.json()["state"] == "ready"

    def test_negotiation_brief_missing_returns_404(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])
        # no brief generated for this scenario

        resp = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="negotiation_brief", format="pdf",
        )
        assert resp.status_code == 404, resp.text
        assert resp.json()["error"]["code"] == "not_found"

    def test_scenario_summary_csv(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])

        resp = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="scenario_summary", format="csv",
        )
        assert resp.status_code == 201, resp.text
        assert resp.json()["state"] == "ready"

    def test_audit_history_csv(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])

        resp = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="audit_history", format="csv",
        )
        assert resp.status_code == 201, resp.text
        body = resp.json()
        assert body["state"] == "ready"

        download = client.get(f"/api/v1/reports/{body['id']}/content", headers=headers)
        assert download.status_code == 200, download.text
        text = download.content.decode("utf-8")
        # the scenario's own creation is itself an audited event
        assert "scenario" in text.lower()


class TestContentDownload:
    def test_sha256_and_headers_match_metadata(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])

        report = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="supplier_comparison", format="xlsx",
        ).json()

        resp = client.get(f"/api/v1/reports/{report['id']}/content", headers=headers)
        assert resp.status_code == 200, resp.text
        assert hashlib.sha256(resp.content).hexdigest() == report["content_sha256"]
        assert len(resp.content) == report["size_bytes"]

        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        disposition = resp.headers["content-disposition"]
        assert "attachment" in disposition
        expected_prefix = f"supplier_comparison-{report['id'].split('-')[0]}"
        assert expected_prefix in disposition
        assert disposition.strip().endswith('.xlsx"')
        assert resp.headers["x-content-type-options"] == "nosniff"

    def test_csv_content_type(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])
        report = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="scenario_summary", format="csv",
        ).json()

        resp = client.get(f"/api/v1/reports/{report['id']}/content", headers=headers)
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("text/csv")

    def test_pdf_content_type(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])
        report = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="cfo_recommendation", format="pdf",
        ).json()

        resp = client.get(f"/api/v1/reports/{report['id']}/content", headers=headers)
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("application/pdf")
        assert resp.content.startswith(b"%PDF")


class TestFormulaInjectionEscape:
    def test_supplier_name_formula_is_escaped_in_csv_and_xlsx(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        """SPEC/task requirement: any cell value starting with =, +, -, @,
        tab, or CR gets a leading apostrophe (app.reports.escape.
        escape_formula_cell), so a supplier name that looks like a
        spreadsheet formula never becomes one."""
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        malicious_name = '=HYPERLINK("http://evil")'
        ctx = _setup_two_supplier_case(
            client, headers, migrated_engine, org_a, supplier_a_name=malicious_name
        )
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])

        csv_report = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="supplier_comparison", format="csv",
        ).json()
        csv_download = client.get(f"/api/v1/reports/{csv_report['id']}/content", headers=headers)
        assert csv_download.status_code == 200
        csv_text = csv_download.content.decode("utf-8")
        assert "'=HYPERLINK" in csv_text
        # the raw, unescaped formula string must never appear
        assert malicious_name not in csv_text

        xlsx_report = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="supplier_comparison", format="xlsx",
        ).json()
        xlsx_download = client.get(
            f"/api/v1/reports/{xlsx_report['id']}/content", headers=headers
        )
        assert xlsx_download.status_code == 200
        wb = load_workbook(io.BytesIO(xlsx_download.content))
        ws = wb.active
        cell_values = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ]
        assert any(v.startswith("'=HYPERLINK") for v in cell_values), cell_values
        assert not any(v == malicious_name for v in cell_values)
        # openpyxl must never have recorded this as a live formula cell
        formula_cells = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if getattr(cell, "data_type", None) == "f"
        ]
        assert formula_cells == []


class TestPurge:
    def test_content_returns_410_after_purge(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])
        report = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="scenario_summary", format="csv",
        ).json()

        with migrated_engine.begin() as conn:
            conn.execute(
                sqltext(
                    "UPDATE generated_reports SET storage_key = NULL, content_sha256 = NULL,"
                    " purged_at = now() WHERE id = :id"
                ),
                {"id": report["id"]},
            )

        resp = client.get(f"/api/v1/reports/{report['id']}/content", headers=headers)
        assert resp.status_code == 410, resp.text
        assert resp.json()["error"]["code"] == "not_found"

        # metadata itself is still readable — the row is kept, only the blob is gone
        meta = client.get(f"/api/v1/reports/{report['id']}", headers=headers)
        assert meta.status_code == 200
        assert meta.json()["purged"] is True


class TestRoleGating:
    def test_viewer_can_list_get_download_but_not_generate(
        self, client: TestClient, org_a: dict[str, Any], migrated_engine: Engine
    ) -> None:
        analyst_headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, analyst_headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, analyst_headers, ctx["rfq"]["id"])
        report = _generate_report(
            client, analyst_headers,
            scenario_id=scenario["id"], report_type="scenario_summary", format="csv",
        ).json()

        viewer_headers = _headers(_login_as(client, org_a, Role.VIEWER))

        listed = client.get("/api/v1/reports", headers=viewer_headers)
        assert listed.status_code == 200, listed.text
        assert any(item["id"] == report["id"] for item in listed.json()["items"])

        got = client.get(f"/api/v1/reports/{report['id']}", headers=viewer_headers)
        assert got.status_code == 200, got.text

        content = client.get(f"/api/v1/reports/{report['id']}/content", headers=viewer_headers)
        assert content.status_code == 200, content.text

        forbidden = _generate_report(
            client, viewer_headers,
            scenario_id=scenario["id"], report_type="scenario_summary", format="csv",
        )
        assert forbidden.status_code == 403, forbidden.text


class TestCrossOrgIsolation:
    def test_org_b_cannot_read_org_a_report_or_content(
        self,
        client: TestClient,
        org_a: dict[str, Any],
        org_b: dict[str, Any],
        migrated_engine: Engine,
    ) -> None:
        headers_a = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers_a, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers_a, ctx["rfq"]["id"])
        report = _generate_report(
            client, headers_a,
            scenario_id=scenario["id"], report_type="scenario_summary", format="csv",
        ).json()

        headers_b = _headers(_login_as(client, org_b, Role.ANALYST))

        got = client.get(f"/api/v1/reports/{report['id']}", headers=headers_b)
        assert got.status_code == 404
        assert got.json()["error"]["code"] == "not_found"

        content = client.get(f"/api/v1/reports/{report['id']}/content", headers=headers_b)
        assert content.status_code == 404
        assert content.json()["error"]["code"] == "not_found"

    def test_org_b_cannot_generate_against_org_a_scenario(
        self,
        client: TestClient,
        org_a: dict[str, Any],
        org_b: dict[str, Any],
        migrated_engine: Engine,
    ) -> None:
        headers_a = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers_a, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers_a, ctx["rfq"]["id"])

        headers_b = _headers(_login_as(client, org_b, Role.ANALYST))
        resp = _generate_report(
            client, headers_b,
            scenario_id=scenario["id"], report_type="scenario_summary", format="csv",
        )
        assert resp.status_code == 404
        assert resp.json()["error"]["code"] == "not_found"


class TestFailedReportErrorHygiene:
    """2026-08 security audit MEDIUM-6: a renderer failure's wire-visible
    error_message must not leak exception detail (SQL, paths, secrets), and
    OrgIsolationViolation must propagate loudly instead of becoming a stored
    `failed` row."""

    def test_error_message_is_generic_and_leak_free(
        self,
        client: TestClient,
        org_a: dict[str, Any],
        migrated_engine: Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])

        def _boom(*args: Any, **kwargs: Any) -> bytes:
            raise RuntimeError("hunter2 at /private/secret/path in SELECT * FROM users")

        monkeypatch.setattr("app.services.report_service.render", _boom)
        resp = _generate_report(
            client, headers,
            scenario_id=scenario["id"], report_type="scenario_summary", format="csv",
        )
        assert resp.status_code == 201, resp.text
        body = resp.json()
        assert body["state"] == "failed"
        message = body["error_message"]
        assert "hunter2" not in message
        assert "/private" not in message
        assert "SELECT" not in message
        assert "RuntimeError" in message  # type name only, plus the report id

    def test_org_isolation_violation_propagates_not_stored(
        self,
        client: TestClient,
        org_a: dict[str, Any],
        migrated_engine: Engine,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        from app.repositories.base import OrgIsolationViolation

        headers = _headers(_login_as(client, org_a, Role.ANALYST))
        ctx = _setup_two_supplier_case(client, headers, migrated_engine, org_a)
        scenario = _create_and_complete_scenario(client, headers, ctx["rfq"]["id"])

        def _violate(*args: Any, **kwargs: Any) -> bytes:
            raise OrgIsolationViolation("cross-org read detected")

        monkeypatch.setattr("app.services.report_service.render", _violate)
        with pytest.raises(OrgIsolationViolation):
            _generate_report(
                client, headers,
                scenario_id=scenario["id"], report_type="scenario_summary", format="csv",
            )

        # The transaction rolled back: no failed row was persisted for it.
        with migrated_engine.connect() as conn:
            count = conn.execute(
                sqltext(
                    "SELECT count(*) FROM generated_reports"
                    " WHERE scenario_id = :sid AND state = 'failed'"
                ),
                {"sid": scenario["id"]},
            ).scalar_one()
        assert count == 0
